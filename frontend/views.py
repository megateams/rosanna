from django.shortcuts import render, redirect, HttpResponse,get_object_or_404
from django.contrib import messages
from .models import *
from django.db.models import Sum
from finance.models import *
from django.contrib.auth.models import User
from django.contrib.auth import login, logout 
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.http import JsonResponse
# from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from datetime import datetime,date
from django.utils.timezone import now
from django.db.models import Max
import openpyxl
from django.core.files.storage import FileSystemStorage
import pandas as pd
from django.db.models import Q  
import bcrypt

salt = bcrypt.gensalt()

def encryptpassword(password):
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), b'$2b$12$QW/1zgrHumeirkSwiM437u')
    return hashed_password

def edit_teacher_class(request):
    if request.method == 'POST':
        teacherid = request.POST.get('teacherid')
        classes = request.POST.getlist('classes')
        that_teacher = Teachers.objects.get(pk=teacherid)
        that_teacher.classes.set(classes)
        # Save the updated teacher
        that_teacher.save()
        messages.success(request,"Teacher classes edited successfully")
        return redirect("Show Teacher", teacherId=teacherid)
    
def edit_teacher_subject(request):
    if request.method == 'POST':
        teacherid = request.POST.get('teacherid')
        subjects = request.POST.getlist('subjects')
        that_teacher = Teachers.objects.get(pk=teacherid)
        that_teacher.subjects.set(subjects)
        # Save the updated teacher
        that_teacher.save()
        messages.success(request,"Teacher classes edited successfully")
        return redirect("Show Teacher", teacherId=teacherid)

def admincheckemail(request , email):
    try:
        Administrators.objects.get(email = email)
        return JsonResponse({'data' : 'Email address already used.'})
    except:
        print("")
    return JsonResponse({'data' : ''})

#register details for the user (admin)
def editadministrator(request):
    if request.method == 'POST':
        adminid = request.POST.get('adminid')
        fullname = request.POST.get('fullname')
        contact = request.POST.get('contact')
        email = request.POST.get('email')
        address = request.POST.get('address')
        salary = request.POST.get('salary')
        gender = request.POST.get('gender')
        qualification = request.POST.get('qualification')
        role = request.POST.get('role')
        
        admin = Administrators.objects.get(pk = adminid)
        
        admin.contact = contact
        admin.email = email
        admin.address = address
        admin.salary = salary
        admin.gender = gender
        admin.qualification = qualification 
        admin.role = role 
        admin.fullname = fullname
        
        admin.save()
        
        messages.success(request , "Administrator Edited Successfully")
        return redirect('adminslist')
    
    admins = Administrators.objects.all()
    return render(request , "frontend/staff/administratorsList.html" , {'admins' : admins})

@login_required
def deleteadmin(request):
    if request.method == 'POST':
        adminid = request.POST.get('adminid')
        deladmin = Administrators.objects.get(id = adminid)
        deladmin.delete()
        messages.success(request , "Administrator Deleted Successfully")
        return redirect("adminslist")
    admins = Administrators.objects.all()
    return render(request , "frontend/staff/administratorsList.html" , {'admins' : admins})

@login_required
def adminslist(request):
    admins = Administrators.objects.all()
    return render(request , 'frontend/staff/administratorsList.html' , {'admins' : admins})

@login_required
def addadmins(request):
    if request.method == 'POST':
        fullname = request.POST.get('fullname')
        gender = request.POST.get('gender')
        address = request.POST.get('address')
        contact = request.POST.get('contact')
        email = request.POST.get('email')
        #profileimage = request.POST.get('profileimage')
        role = request.POST.get('role')
        qualification = request.POST.get('qualification')
        bankaccnum = request.POST.get('bankaccnum')
        salary = request.POST.get('salary')
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        try:
            adminexist = Administrators.objects.get(username = username)
            messages.error(request , "Administrator already exists.")
            return render(request , 'frontend/staff/administratorsAdd.html')
        except:
            Administrators.objects.create(
                fullname = fullname ,
                gender = gender ,
                address = address ,
                contact = contact ,
                email = email ,
                role = role ,
                qualification = qualification ,
                bankaccnum =  bankaccnum ,
                salary = salary ,
                username = username ,
                password = encryptpassword(password) ,
            )
            
            Administrators.save
            # User.objects.create_user(
            #     username = username ,
            #     password = password ,
            #     email = email ,
            #     is_staff = True ,
            #     is_superuser = True ,
            # )
            
            messages.success(request , "Administrator Created Successfully")
            return redirect('adminslist') 

    return render(request , 'frontend/staff/administratorsAdd.html')

def register(request):
    if request.method == "POST":
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("login")  # Redirect to the login page after successful registration
    else:
        form = UserCreationForm()
    return render(request, "frontend/registration.html", {"form": form})

# login views for the admin user

def user_login(request):
    if request.method == "POST":
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            # Log the user in
            user = form.get_user()
            login(request, user)
            request.session.pop('logged_out', None) # clear logout session variable in login
            return redirect("Dashboard")
        else:
            messages.error(request, 'Invalid username or password.')
    else:
        form = AuthenticationForm()
    return render(request, "frontend/login.html")

@login_required
def user_logout(request):
    logout(request)
    messages.success(request, "You have successfully logged out")
    request.session['logged_out'] = True # Set the session variable to True
    return redirect("Admin Login") # Redirect to the login page after logout
# Create your views here.
# creating views for dashboard

@login_required
def home(request):
    if request.session.get('logged_out', False):
        messages.warning(request, "You need to login to access the dashboard")
        return redirect("login")

    boys_count = Student.objects.filter(gender='m').count()
    girls_count = Student.objects.filter(gender='f').count()

    teacher_count = Teachers.objects.count()
    student_count = Student.objects.count()
    support_staff_count = Supportstaff.objects.count()

    classes = Schoolclasses.objects.all()

    # Calculate the number of boys and girls in each class and store in lists
    boys_count_by_class = []
    girls_count_by_class = []
    class_names = []
    for class_obj in classes:
        b_count = Student.objects.filter(stdclass=class_obj, gender='m').count()
        g_count = Student.objects.filter(stdclass=class_obj, gender='f').count()
        boys_count_by_class.append(b_count)
        girls_count_by_class.append(g_count)
        class_names.append(class_obj.classname)
        
    context = {
        'boys_count' : boys_count,
        'girls_count': girls_count,
        'teacher_count': teacher_count,
        'student_count': student_count,
        'support_staff_count': support_staff_count,

        'boys_count_by_class': boys_count_by_class,
        'girls_count_by_class': girls_count_by_class,
        'class_names': class_names,
        'term_data' : Term.objects.filter(status=1),
    }
    return render(request,'frontend/dashboard.html',context)

# students views
@login_required
def studentsList(request):
    #retrieve all the selected students data from the database
    selected_students =Student.objects.all()
    classes = Schoolclasses.objects.all()

    context = {
      'students': selected_students,
       'classes': classes
    }
    #pass the data to template for rendering
    return render(request, 'frontend/student/studentsList.html',context)

@login_required
def studentsAdd(request):
    return render(request, 'frontend/student/studentsAdd.html',{'classes': Schoolclasses.objects.all()})

@login_required
def student_by_class(request, class_id):
    try:
        # Fetch the class based on the class_id
        selected_class = Schoolclasses.objects.get(classid=class_id)
        print(selected_class)
        # Fetch students in the selected class
        students = Student.objects.filter(stdclass=selected_class)
        classes = Schoolclasses.objects.all()
        return render(request, 'frontend/student/student_by_class.html', {'students': students, 'selected_class': selected_class, 'classes': classes})
    except Exception as e:
        return render(request, 'frontend/student/studentsList.html', {'error_message': str(e)})

# teachers views
@login_required
def teacherAdd(request):
    classes = Schoolclasses.objects.all()
    subjects = Subjects.objects.all()
    return render(request,'frontend/staff/teacherAdd.html', {'classes': classes, 'subjects':subjects})

@login_required
def teacherList(request):
    teachers = Teachers.objects.all()
    return render(request,'frontend/staff/teacherList.html', {'teachers': teachers})
# teachers views

# users views
@login_required
def addUsers(request):
    return render(request,'frontend/users/addUsers.html')

@login_required
def usersList(request):
    return render(request,'frontend/users/usersList.html')
# users views
@login_required
def get_subjects(request, class_id):
    class_obj = get_object_or_404(Schoolclasses, pk=class_id)
    subjects = class_obj.subjects.all().values('subjectid', 'subjectname')
    return JsonResponse(list(subjects), safe=False)


@login_required
def subjectList(request):
    subjects = Subjects.objects.all()
    teachers = Teachers.objects.all()
    # print(teachers)
    return render(request,'frontend/subjects/subjectList.html',{'subjects':subjects, 'teachers':teachers})
# classes views

@login_required
def showStudent(request):
    return render(request, 'frontend/student/showStudent.html')

def addMarks(request):
    # frontend/views.py
    if request.method == 'POST':
        std_id = request.POST['student_name']
        class_id = request.POST['class_name']
        subject_id = request.POST['subject']
        score = request.POST['score']

        # std_obj = Student.objects.get(pk=std_id)
        class_obj = Schoolclasses.objects.get(pk=class_id)
        subject_obj = Subjects.objects.get(pk=subject_id)

        Mark.objects.create(
            class_name=class_obj,
            student_name=std_id,
            subject=subject_obj,
            marks_obtained=score
        )
        Mark.save
        messages.success(request, 'Data successfully added!')
        return redirect('Add Marks')  # Redirect to the same form page after successful form submission
    else:
        selected_students =Student.objects.all()
        # messages.success(request, 'Data not added!')
        return render(request,'frontend/marks/addMarks.html',{'classes': Schoolclasses.objects.all() ,'students': selected_students})


@login_required
def marksList(request):
    classes = Schoolclasses.objects.all()
    return render(request,'frontend/marks/marksList.html', {'marks': Mark.objects.all(), 'classes':classes})
# marks views


@login_required
def view_marks(request, class_id):
    # Retrieve the class and teacher objects based on the provided IDs
    schoolclass = get_object_or_404(Schoolclasses, classid=class_id)

    # Get the students for this class taught by the teacher
    students = Student.objects.filter(stdclass=schoolclass)

    # Get all subjects
    subjects = Subjects.objects.filter(schoolclasses=schoolclass)
    subjects_count = Subjects.objects.filter(schoolclasses=schoolclass).count()

    mark_types = Mark.MARK_TYPES
    term_data = Term.objects.get(status=1)
    # Create a dictionary to hold the total and average marks for each subject for each student
    subjects_marks_data = {}
    subjects_total_marks = {subject: 0 for subject in subjects}
    subjects_marks_count = {subject: 0 for subject in subjects}

    for student in students:
        student_subjects_data = []
        total_average_marks = 0  # Initialize total_average_marks for each student
        for subject in subjects:
            total_marks = 0
            marks_count = 0
            marks = Mark.objects.filter(class_name=schoolclass, student_name=student.childname, subject=subject).exclude(mark_type='Test')

            for mark in marks:
                total_marks += mark.marks_obtained
                marks_count += 1
                subjects_total_marks[subject] += mark.marks_obtained
                subjects_marks_count[subject] += 1

            average_marks = total_marks / 3 if marks_count > 0 else 0
            total_average_marks += int(average_marks)
            final_average = total_average_marks / subjects_count if marks_count > 0 else 0
              # Accumulate average marks for the student
            student_subjects_data.append({
                'subject': subject,
                'total_marks': total_marks,
                'average_marks': average_marks,
            })
        subjects_marks_data[student] = student_subjects_data
        student.final_average = final_average
        student.total_average_marks = total_average_marks  # Store total average marks for the student

    # Sort the students based on their total marks in descending order
    students = sorted(students, key=lambda student: student.total_average_marks, reverse=True)

    # Assign ranks to students based on their position in the sorted list
    for rank, student in enumerate(students, start=1):
        student.rank = rank

    return render(request,"frontend/marks/view_marks.html",{
        'schoolclass': schoolclass,
        'students': students,
        'subjects': subjects,
        'mark_types': mark_types,
        'subjects_marks_data': subjects_marks_data,
    })

@login_required
def view_marks_by_type(request, class_id):
    schoolclass = Schoolclasses.objects.get(classid=class_id)
    mark_type = request.GET.get('marktype')  # Get the mark type from the query parameter
    
    students = Student.objects.filter(stdclass=schoolclass)
    subjects = Subjects.objects.filter(schoolclasses=schoolclass)
    mark_types = Mark.MARK_TYPES
    student_marks = {}
    for student in students:
        marks = Mark.objects.filter(class_name=schoolclass, student_name=student.childname, mark_type=mark_type)
        student_marks[student] = {subject: marks.filter(subject=subject).first() for subject in subjects}

    # Calculate total marks and average marks for each student
    for student, marks_dict in student_marks.items():
        total_marks = 0
        total_subjects = 0
        for marks in marks_dict.values():
            if marks:
                total_marks += marks.marks_obtained
                total_subjects += 1
        if total_subjects > 0:
            student.total_marks = total_marks
            student.average_marks = total_marks / total_subjects
        else:
            student.total_marks = 0
            student.average_marks = 0
    return render(request, 'frontend/marks/view_marks_by_type.html', {
        'schoolclass': schoolclass,
        'students': students,
        'subjects': subjects,
        'student_marks': student_marks,
        'mark_types' :mark_types,
        'mark_type' :mark_type,
    })


@login_required
def showStudent(request,stdnumber):
    student = Student.objects.get(stdnumber = stdnumber)
    classes = Schoolclasses.objects.all()
    return render(request, 'frontend/student/showStudent.html', {'student': student, 'classes':classes})
# students views

# edit profile image
def edit_std_image(request):
    if request.method == "POST":
        stdnumber = request.POST.get("stdnumber")
        profile_image = request.FILES.get('profile_image')

        student = Student.objects.get(stdnumber=stdnumber)

        if profile_image:
            fs = FileSystemStorage()
            image_filename = fs.save(profile_image.name, profile_image)
            student.profile_image = image_filename

            student.save()
            messages.success(request, "Image edited successfully")
            return redirect("Student details", stdnumber=stdnumber)
        
        else:
            messages.success(request, "Error!!!")
            return redirect("Student details", stdnumber=stdnumber)  

# edit profile image
def edit_tr_image(request):
    if request.method == "POST":
        teacherid = request.POST.get("teacherid")
        profile_image = request.FILES.get('profile_image')

        teacher = Teachers.objects.get(pk=teacherid)

        if profile_image:
            fs = FileSystemStorage()
            image_filename = fs.save(profile_image.name, profile_image)
            teacher.profile_image = image_filename

            teacher.save()
            messages.success(request, "Image edited successfully")
            return redirect("Show Teacher", teacherId=teacherid)
        
        else:
            messages.success(request, "Error!!!")
            return redirect("Show Teacher", teacherId=teacherid)  



# edit student
@login_required
def edit_student(request):
    if request.method == "POST":
        stdnumber = request.POST.get("stdnumber")
        regdate = request.POST.get('regdate')
        childname = request.POST.get('childname')
        gender = request.POST.get('gender')
        dob = request.POST.get('dob')
        address = request.POST.get('address')
        house = request.POST.get('house')
        fathername = request.POST.get('fathername')
        fcontact = request.POST.get('fcontact')
        foccupation = request.POST.get('foccupation')
        mothername = request.POST.get('mothername')
        mcontact = request.POST.get('mcontact')
        moccupation = request.POST.get('moccupation')
        livingwith = request.POST.get('livingwith')
        guardianname = request.POST.get('guardianname')
        gcontact = request.POST.get('gcontact')
        username = request.POST.get("username")

        student = Student.objects.get(pk=stdnumber)
        student.regdate = regdate
        student.childname = childname
        student.gender = gender
        student.dob = dob
        student.address = address
        student.house = house
        student.fathername = fathername
        student.fcontact = fcontact
        student.foccupation = foccupation
        student.mothername = mothername
        student.mcontact = mcontact
        student.moccupation = moccupation
        student.livingwith = livingwith
        student.guardianname = guardianname
        student.gcontact = gcontact
        student.username = username

        student.save()
        messages.success(request, 'Student edited successfully')
        return redirect("Student details", stdnumber=stdnumber)

def delete_student(request):
    if request.method == "POST":
        stdnumber = request.POST.get("stdnumber")

        student = Student.objects.get(pk=stdnumber)
        student.delete()
        messages.success(request, "Student has been deleted")

        return redirect("Students List")

# support staff views
@login_required
def supportstaffAdd(request):
      return render(request, 'frontend/staff/supportstaffAdd.html')

@login_required
def supportstaffreg(request):
    if request.method == 'POST':
        supportstaffnames = request.POST.get('supportstaffnames')
        dob = request.POST.get('dob')

        # Check if a support staff with the same name and date of birth already exists
        if Supportstaff.objects.filter(Q(supportstaffnames=supportstaffnames) & Q(dob=dob)).exists():
            messages.error(request, 'A support staff with the same name and date of birth already exists.')
            return redirect("AddSupportstaff")  # Redirect back to the form page

        profile_image = request.FILES.get('profile_image')
        last_supportstaff = Supportstaff.objects.order_by('-supportstaffid').first()

        if last_supportstaff:
            default_supportstaffid = 'RSS{:04}'.format(int(last_supportstaff.supportstaffid[3:]) + 1) 
        else:
            default_supportstaffid = 'RSS0001' 

        # Capture the current date
        current_date = date.today()


        # Retrieve other form fields
        gender = request.POST.get('gender')
        contact = request.POST.get('contact')
        email = request.POST.get('email')
        address = request.POST.get('address')
        # joiningdate = request.POST.get('joiningdate')
        qualification = request.POST.get('qualification')
        position = request.POST.get('position')
        salary = request.POST.get('salary')
        bankaccnum = request.POST.get('bankaccnum')

        # Create and save the SupportStaff object to the database
        staff = Supportstaff.objects.create(
            supportstaffid=default_supportstaffid,
            supportstaffnames=supportstaffnames,
            dob=dob,
            gender=gender,
            contact=contact,
            email=email,
            address=address,
            joiningdate=current_date,
            qualification=qualification,
            position=position,
            salary=salary,
            bankaccnum=bankaccnum,
            # Add other fields as needed
        )

        if profile_image:
            fs = FileSystemStorage()
            image_filename = fs.save(profile_image.name, profile_image)
            staff.profile_image = image_filename
        staff.save()

        # Display a success message to the user
        messages.success(request, 'Support staff added successfully!')

        # Redirect the user to a different page (e.g., supportstafflist page)
        return redirect('AddSupportstaff')

def edit_sstaff_image(request):
    if request.method == "POST":
        supportstaffid = request.POST.get("supportstaffid")
        profile_image = request.FILES.get('profile_image')
        
        staff = Supportstaff.objects.get(pk=supportstaffid)

        if profile_image:
            fs = FileSystemStorage()
            image_filename = fs.save(profile_image.name, profile_image)
            staff.profile_image = image_filename

            staff.save()
            messages.success(request, "Image edited successfully")
            return redirect("show supportstaff", supportstaffid=supportstaffid)
        
        else:
            messages.success(request, "Error!!!")
            return redirect("show supportstaff", supportstaffid=supportstaffid)  

@login_required
def addsubjectsform(request):
    return render(request, 'frontend/subjects/addSubject.html')

def addsubject(request):
    if request.method == 'POST':
        subject_name = request.POST.get('subjectname')
        # classlevel = request.POST.get('classlevel')
        # subjectheads = request.POST.get('subjecthead')

        # Check if the subject already exists
        if Subjects.objects.filter(subjectname=subject_name).exists():
            messages.error(request, 'Subject with this name already exists.')
        else:
            # Find the last subject ID to determine the next one
            last_subject = Subjects.objects.order_by('-subjectid').first()
            if last_subject:
                last_subject_id = last_subject.subjectid
                next_subject_number = int(last_subject_id[3:]) + 1
            else:
                next_subject_number = 1

            # Generate the new subject ID
            subjectids = 'SUB{:02d}'.format(next_subject_number)
            
            Subjects.objects.create(
                subjectname=subject_name,
                subjectid=subjectids,
                # classlevel=classlevel,
                # subjecthead=subjectheads
            )
            Subjects.save
            messages.success(request, 'Subject successfully added!')
    return redirect("addsubjectsform")
    # return render(request , 'frontend/academics/subjects.html' , {'subjects' : Subjects.objects.all()})

# assign subject head
def assign_subjecthead(request):
    if request.method == 'POST':
        subject_id = request.POST.get("subject_id")
        subject_head = request.POST.get("subject_head")
        
        # Check if the teacher is already assigned as a subject head for another subject
        if Subjects.objects.filter(subjecthead=subject_head).exclude(pk=subject_id).exists():
            messages.error(request, f'Teacher {subject_head} is already assigned as a subject head to another subject.')
        else:
            that_subject = Subjects.objects.get(pk=subject_id)
            that_subject.subjecthead = subject_head
            that_subject.save()
            messages.success(request, f'Teacher {subject_head} assigned successfully to {that_subject.subjectname}.')
    
    return redirect("subjectList")


# edit subject view
def edit_subject(request):
    if request.method == 'POST':
        subject_id = request.POST.get("subject_id")
        subjectname = request.POST.get("subjectname")
        classlevel = request.POST.get("classlevel")
        subject_head = request.POST.get("subject_head")
        that_subject = Subjects.objects.get(pk=subject_id)

        that_subject.subjectname = subjectname
        that_subject.classlevel = classlevel
        that_subject.subjecthead = subject_head
        that_subject.save()
        messages.success(request,"Subject edited successfully")
        return redirect("subjectList")

# delete subject view
def delete_subject(request):
    if request.method == 'POST':
        subject_id = request.POST.get("subject_id")
        that_subject = Subjects.objects.get(pk=subject_id)
        that_subject.delete()
        messages.success(request,"Subject deleted successfully")
        return redirect("subjectList")

@login_required
def supportstaffList(request):
    # Retrieve all support staff data from the database
    all_support_staff = Supportstaff.objects.all()
    # Pass the data to the template for rendering
    return render(request, 'frontend/staff/supportstaffList.html', {'support_staff': all_support_staff})

@login_required   
def showteacher(request,teacherId):
    teacher = Teachers.objects.get(teacherid=teacherId)
    classes = Schoolclasses.objects.all()
    subjects = Subjects.objects.all()
    return render(request, 'frontend/staff/showteacher.html',{'teacher': teacher, 'classes':classes, 'subjects':subjects})


@login_required
def showsupportstaff(request,supportstaffid):
    supportstaffs = Supportstaff.objects.filter(supportstaffid=supportstaffid)
    return render(request, 'frontend/staff/showsupportstaff.html',{'supportstaffs': supportstaffs})
# support staff views

#students registration views
def studentReg(request):
    if(request.method == 'POST'):
        dob = request.POST.get('dob')
        childname = request.POST.get('childname')

        # Check if a student with the same name and date of birth already exists
        if Student.objects.filter(Q(childname=childname) & Q(dob=dob)).exists():
            messages.error(request, 'A student with the same name and date of birth already exists.')
            return redirect("AddStudents")  # Redirect back to the form page

        # capturing the profile image from the form
        profile_image = request.FILES.get('profile_image')
        
        # Find the maximum stdnumber in the database
        max_stdnumber = Student.objects.aggregate(Max('stdnumber'))['stdnumber__max']

        # Generate the next stdnumber
        if max_stdnumber:
            new_stdnumber = 'STD{:03d}'.format(int(max_stdnumber[3:]) + 1)
        else:
            new_stdnumber = 'STD001'
        default_password = "123456"

        # Capture the current date
        current_date = date.today()


        # regdate = request.POST.get('regdate')
        childname = request.POST.get('childname')
        class_id = request.POST['stdclass']
        selected_class = Schoolclasses.objects.get(pk=class_id)
        gender = request.POST.get('gender')
        dob = request.POST.get('dob')
        address = request.POST.get('address')
        house = request.POST.get('house')
        username = request.POST.get('username')
        fathername = request.POST.get('fathername')
        fcontact = request.POST.get('fcontact')
        foccupation = request.POST.get('foccupation')
        mothername = request.POST.get('mothername')
        mcontact = request.POST.get('mcontact')
        moccupation = request.POST.get('moccupation')
        livingwith = request.POST.get('livingwith')
        guardianname = request.POST.get('guardianname')
        gcontact = request.POST.get('gcontact')
        
        
        student =Student.objects.create(
            stdnumber =new_stdnumber,            
            regdate = current_date ,
            childname = childname ,
            stdclass=selected_class,
            gender = gender ,
            dob = dob ,
            address = address ,
            house = house ,
            username = username ,
            password = encryptpassword(default_password) ,
            fathername = fathername ,
            fcontact = fcontact ,
            foccupation = foccupation ,
            mothername = mothername , 
            mcontact = mcontact ,
            moccupation = moccupation ,
            livingwith = livingwith ,
            guardianname = guardianname ,
            gcontact = gcontact
        )

        if profile_image:
            fs = FileSystemStorage()
            image_filename = fs.save(profile_image.name, profile_image)
            student.profile_image = image_filename

        student.save ()
        messages.success(request, 'Data successfully added!')
        return redirect("AddStudents")
        # return HttpResponse('Registration Successful')

def get_students_by_class(request, class_id):
    try:
        selected_class = Schoolclasses.objects.get(pk=class_id)
        students = Student.objects.filter(stdclass=selected_class)
        data = {
            'students': list(students.values('pk', 'childname')),
            # Add other fields you want to include in the JSON response
        }
        return JsonResponse(data)
    except Schoolclasses.DoesNotExist:
        data = {'error': 'Class not found'}
        return JsonResponse(data, status=404)

def subjects(request):
    if request.method == 'POST':
        subjectnames = request.POST.get('subjectname')
        subjectids = request.POST.get('subjectid')
        classlevels = request.POST.get('classlevel')
        subjectheads = request.POST.get('subjecthead')
        
        try:
            Subjects.objects.get(sugjectname = subjectnames , classlevels = classlevels)
            messages.error(request , 'Subject Already Exists')
        except:
            Subjects.objects.create(
                subjectname = subjectnames , 
                subjectid = subjectids , 
                classlevel = classlevels , 
                subjecthead = subjectheads
            )
        
            Subjects.save
            messages.success(request , "Subject Added Succesfully")
    return HttpResponse(subjectnames)

@login_required
def showclasses(request):
    classes = Schoolclasses.objects.all()
    # Retrieve the teacher names based on teacherid matching classname
    teachers = Teachers.objects.all()
    all_subjects = Subjects.objects.all()
    return render(request, 'frontend/academics/showclasses.html', {'classes': classes, 'teachers': teachers,'all_subjects':all_subjects})

@login_required
def addclasses(request):
    subjects = Subjects.objects.all()
    teachers = Teachers.objects.all()
    return render(request , 'frontend/academics/addclasses.html' , {'subjects':subjects, 'teachers':teachers})

#add classes view 
def schoolclasses(request):
    if request.method == 'POST':
        classname = request.POST['class_name']
        subject_names = request.POST.getlist('subjects')
        # class_level = request.POST['class_level']

        # Check if the class already exists
        if Schoolclasses.objects.filter(classname=classname).exists():
            messages.error(request, 'Class with this name already exists.')
            return redirect("AddClasses")  # Redirect back to the form page

        newclass = Schoolclasses.objects.create(classname=classname)

        for subject_name in subject_names:
            subject, created = Subjects.objects.get_or_create(subjectname=subject_name)
            newclass.subjects.add(subject)

        messages.success(request, 'Class successfully added!')
        return redirect("AddClasses")


# edit class view
def edit_class(request):
    if request.method == 'POST':
        classid = request.POST.get("classid")
        classname = request.POST.get("classname")
        classteacher = request.POST.get("classteacher")

        # Check if the teacher is already a classteacher for another class
        if Schoolclasses.objects.filter(classteacher=classteacher).exclude(pk=classid).exists():
            messages.error(request, f'{classteacher} is already a classteacher for another class.')
        else:
            that_class = Schoolclasses.objects.get(pk=classid)
            that_class.classname = classname
            that_class.classteacher = classteacher
            that_class.save()
            messages.success(request, "Class edited successfully")

    return redirect("showclasses")

# modal for deleting class
def delete_class(request):
    if request.method == 'POST':
        classid = request.POST.get("classid")
        that_class = Schoolclasses.objects.get(pk=classid)
        that_class.delete()
        messages.success(request,"Class deleted successfully")
        return redirect("showclasses")

#Export the data in excel in the students model
# import openpyxl
# from django.http import HttpResponse

def export_to_excel(request):
    try:
        students = Student.objects.all()

        # Create a new workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write headers to the worksheet
        headers = [
            'Student Number', 'Child Name', 'Gender', 'Date of Birth', 'Address', 'House', 'Class',
            'Registration Date', "Father's Name", "Father's Contact", "Father's Occupation",
            "Mother's Name", "Mother's Contact", "Mother's Occupation", 'Living With',
            "Guardian's Name", "Guardian's Contact"
        ]
        ws.append(headers)

        # Set column widths for all columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 20
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 20

        # Write student data to the worksheet
        for student in students:
            ws.append([
                student.stdnumber, student.childname, student.gender, student.dob,
                student.address, student.house, student.stdclass.classname, student.regdate,
                student.fathername, student.fcontact, student.foccupation,
                student.mothername, student.mcontact, student.moccupation,
                student.livingwith, student.guardianname, student.gcontact
            ])

        # Create an HttpResponse with the Excel file
        filename = 'students_data.xlsx'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save the workbook to the response
        wb.save(response)
        messages.success(request, 'Data successfully exported to Excel.')
        return response

    except Schoolclasses.DoesNotExist:
        return HttpResponse('Class not found', status=404)
    


def export_students_by_class(request, class_id):
    try:
        selected_class = Schoolclasses.objects.get(pk=class_id)
        students = Student.objects.filter(stdclass=selected_class)

        # Create a new workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write headers to the worksheet
        headers = [
            'Student Number', 'Child Name', 'Gender', 'Date of Birth', 'Address', 'House', 'Class',
            'Registration Date', "Father's Name", "Father's Contact", "Father's Occupation",
            "Mother's Name", "Mother's Contact", "Mother's Occupation", 'Living With',
            "Guardian's Name", "Guardian's Contact"
        ]
        ws.append(headers)

        # Set column widths for all columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 20
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 20

        # Write student data to the worksheet
        for student in students:
            ws.append([
                student.stdnumber, student.childname, student.gender, student.dob,
                student.address, student.house, student.stdclass.classname, student.regdate,
                student.fathername, student.fcontact, student.foccupation,
                student.mothername, student.mcontact, student.moccupation,
                student.livingwith, student.guardianname, student.gcontact
            ])

        # Create an HttpResponse with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="students_in_{selected_class.classname}.xlsx"'
        wb.save(response)
        return response

    except Schoolclasses.DoesNotExist:
        return HttpResponse('Class not found', status=404)

#Export the data in excel in the support staff model
def support_staff_export_to_excel(request):
    #fetch all the data from the database
        data =Supportstaff.objects.all().values()
        #create new workbook and add in a worksheet
        wb =openpyxl.Workbook()
        ws =wb.active
        
        #write field names to the worksheet as headers
        field_names = Supportstaff._meta.get_fields()
        header_row = [field.name for field in field_names]
        ws.append(header_row)

        # Set column widths for all columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        
        #write data to the worksheet
        for row_data in data:
            ws.append(list(row_data.values()))
            
        # Set the filename and content type for the response
        filename = 'supportstaff_data.xlsx'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        #save the workboot to the response
        wb.save(response)
        return response


#Export the data in excel in the Teacher's model
def teacher_export_to_excel(request):
    #fetch all the data from the database
        data =Teachers.objects.all().values_list(
            'teacherid', 'teachernames', 'dob', 'gender', 'contact', 'email', 'address', 'classes', 'joiningdate','subjects', 'qualification'
            )
        
        #create new workbook and add in a worksheet
        wb =openpyxl.Workbook()
        ws =wb.active
        
        #write data to the worksheet
        ws.append(['Teacher ID', 'Name', 'DOB', 'Gender', 'Contact', 'Email', 'Address', 'Classes Taught', 'Date Joined','Subjects', 'Qualification'])
        
        # Set column widths for all columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 20

        for row_data in data:
            ws.append(row_data)
            
        # Set the filename and content type for the response
        filename = 'teacher\'s_data.xlsx'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        #save the workboot to the response
        wb.save(response)
        
        return response

def admins_export_to_excel(request):
    # Fetch all the data from the database
    data = Administrators.objects.all().values()

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define the fields to export (excluding 'profileimage' and 'password')
    fields_to_export = [
        'fullname', 'gender', 'address', 'contact', 'email',
        'role', 'qualification', 'salary', 'bankaccnum'
    ]

    # Write field names to the worksheet as headers
    ws.append(fields_to_export)

    # Set column widths for all columns
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15

    # Write data to the worksheet
    for row_data in data:
        # Extract values only for the fields to export
        row_values = [row_data[field] for field in fields_to_export]
        ws.append(row_values)

    # Set the filename and content type for the response
    filename = 'administrators_data.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response
    
    
def teachers(request):
    if request.method == 'POST':
        teachernames = request.POST.get('teachernames')
        dob = request.POST.get('dob')

        # Check if a teacher with the same name and date of birth already exists
        if Teachers.objects.filter(Q(teachernames=teachernames) & Q(dob=dob)).exists():
            messages.error(request, 'A teacher with the same name and date of birth already exists.')
            return redirect("Add Teacher")  # Redirect back to the form page

        profile_image = request.FILES.get('profile_image')
        last_teacher = Teachers.objects.order_by('-teacherid').first()

        # Generate default ID and password
        default_teacherid = 'RT{:04}'.format(int(last_teacher.teacherid[2:]) + 1) if last_teacher else 'RT0001'
        default_password = "123456"

        # Capture the current date
        current_date = date.today()


        # teacherid = request.POST.get('teacherid')
        teachernames = request.POST.get('teachernames')
        gender = request.POST.get('gender')
        dob = request.POST.get('dob')
        contact = request.POST.get('contact')
        email = request.POST.get('email')
        address = request.POST.get('address')
        # joiningdate = request.POST.get('joiningdate')
        classes = request.POST.getlist('classes')  # Get a list of selected classes
        subjects = request.POST.getlist('subjects')  # Get a list of selected subjects
        qualification = request.POST.get('qualification')
        username = request.POST.get('username')
        salary = request.POST.get('salary')
        bankaccnum = request.POST.get('bankaccnum')
        # password = request.POST.get('password')


        # Create and save the Teacher object to the database
        teacher = Teachers.objects.create(
            teacherid=default_teacherid,
            teachernames=teachernames,
            gender=gender,
            dob=dob,
            contact=contact,
            email=email,
            address=address,
            joiningdate=current_date,
            qualification=qualification,
            salary=salary,
            bankaccnum=bankaccnum,
            username=username,
            password=encryptpassword(default_password)
        )
        if profile_image:
            fs = FileSystemStorage()
            image_filename = fs.save(profile_image.name, profile_image)
            teacher.profile_image = image_filename
        teacher.save()

        # Add the selected classes and subjects to the teacher object
        teacher.classes.set(classes)
        teacher.subjects.set(subjects)

        # Display a success message to the user
        messages.success(request, 'Teacher added successfully!')

        # Redirect the user to a different page (e.g., teacherlist page)
        return redirect('Add Teacher')

# edit teacher view
def edit_teacher(request):
    if request.method == 'POST':
        teacherid = request.POST.get("teacherid")
        teachernames = request.POST.get('teachernames')
        gender = request.POST.get('gender')
        
        contact = request.POST.get('contact')
        email = request.POST.get('email')
        address = request.POST.get('address')
        classes = request.POST.getlist('classes')  # Get a list of selected classes
        subjects = request.POST.getlist('subjects')  # Get a list of selected subjects
        position = request.POST.get('position')
        qualification = request.POST.get('qualification')
        username = request.POST.get('username')
        salary = request.POST.get('salary')
        bankaccnum = request.POST.get('bankaccnum')

        that_teacher = Teachers.objects.get(pk=teacherid)
        print(teacherid)

         # Update teacher attributes
        that_teacher.teachernames = teachernames
        that_teacher.gender = gender
        that_teacher.contact = contact
        that_teacher.email = email
        that_teacher.address = address
        that_teacher.position = position
        that_teacher.qualification = qualification
        that_teacher.username = username
        that_teacher.salary = salary
        that_teacher.bankaccnum = bankaccnum

        # Update related fields (classes and subjects)
        that_teacher.classes.set(classes)
        that_teacher.subjects.set(subjects)

        # Save the updated teacher
        that_teacher.save()

        messages.success(request,"Teacher edited successfully")
        return redirect("Show Teacher", teacherId=teacherid)

def edit_class_subjects(request):
    # Get all available subjects
    all_subjects = Subjects.objects.all()

    if request.method == 'POST':
        # Get the list of selected subjects from the POST request
        classid = request.POST.get("classid")
        selected_subjects = request.POST.getlist('subjects[]')

        # Update the subjects for the class
        class_instance = get_object_or_404(Schoolclasses, pk=classid)
        class_instance.subjects.set(selected_subjects)

        # Save the updated class
        class_instance.save()

        return redirect('showclasses')

    # Handle GET requests or other cases if needed
    return render(request, 'frontend/academics/showclasses.html', {'class_instance': class_instance, 'all_subjects': all_subjects})

@login_required
def supportstaffList(request):
    # Retrieve all support staff data from the database
    all_support_staff = Supportstaff.objects.all()
    # Pass the data to the template for rendering
    return render(request, 'frontend/staff/supportstaffList.html', {'support_staff': all_support_staff})

# Accounting
@login_required
def feesstructure(request):
    feesstructure = Feesstructure.objects.all()
    return render(request, 'frontend/accounting/feesstructure.html',{"feesstructure": feesstructure})

@login_required
def fees(request):
    total_amount = Fees.objects.aggregate(Sum('amount'))['amount__sum']
    fees_list = Fees.objects.all()
    classes = Schoolclasses.objects.all()
    context = {
        'fees_list': fees_list,
        'total_amount': total_amount,
        'classes': classes,
    }
    return render(request, 'frontend/accounting/fees.html',context)

@login_required
def fees_by_class(request, class_id):
    classes = Schoolclasses.objects.all()
    
    try:
        # Fetch the selected class based on the class_id
        selected_class = Schoolclasses.objects.get(classid=class_id)
        
        # Fetch fees records for students in the selected class
        fees_list = Fees.objects.filter(studentclass=selected_class.classname)
        
        # Retrieve the class fees from the Fees model
        classfees = Fees.objects.filter(studentclass=selected_class.classname).first()
        classfees = classfees.classfees if classfees else 0  # Default value if class fees are not found
    except Schoolclasses.DoesNotExist:
        selected_class = None
        fees_list = []
        classfees = 0  # Default value if class is not found or fees not set
    
    return render(request, 'frontend/accounting/fees_by_class.html', {
        'fees_list': fees_list,
        'classes': classes,
        'selected_class': selected_class,
        'classfees': classfees,  # Pass the class fees to the template
    })

@login_required
def teacherspayments(request):
    return render(request, 'frontend/accounting/teacherspayments.html')

@login_required
def supportstaffpayments(request):
    return render(request, 'frontend/accounting/supportstaffpayments.html')

@login_required
def expenses(request):
    total_amount_paid = ExpenseRecord.objects.aggregate(Sum('amountpaid'))['amountpaid__sum']
    expenses = ExpenseRecord.objects.all()
    context = {
        'expenses': expenses,
        'total_amount_paid': total_amount_paid,
        }
    return render(request, 'frontend/accounting/expenses.html',context)

@login_required
def financeteacherpaymentsList(request):
    teachers = Teacherspayment.objects.all()
    return render(request,'frontend/accounting/teacherspayments.html' , {'teachers':teachers})

@login_required
def supportstaffpaymentsList(request):
    supportstaffinfo = Supportstaffpayment.objects.all()
    return render(request,'frontend/accounting/supportstaffpayments.html' , {'supportstaffdata':supportstaffinfo})


def export_subjects_to_excel(request):
    data = Subjects.objects.all().values_list(
        'subjectid', 'subjectname', 'subjecthead'
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append([
        'Subject ID', 'Subject Name', 'Subject Head'
    ])

    for row_data in data:
        ws.append(row_data)

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    filename = 'subjects.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response

def export_classes_to_excel(request):
    data = Schoolclasses.objects.all().values_list(
        'classname', 'subjects', 'classteacher'
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append([
        'Class Name', 'Subjects', 'Class Teacher'
    ])

    for row_data in data:
        ws.append(row_data)

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['D'].width = 15

    filename = 'classes.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response

# settings views
@login_required
def settings(request):
    if request.method == "POST":
        current_term = request.POST.get("term")
        current_year = request.POST.get("year")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        # Check if the end_date of the previous term has been reached
        previous_term = Term.objects.filter(status=1).first()
        if previous_term and date.today() < previous_term.end_date:
            messages.error(request, "Cannot add a new term before the current term ends.")
            return redirect("settings")

        # First, set the status of any existing terms to expired
        Term.objects.update(status=0)
        
        # Create a new Term object and set its status to current
        new_term = Term.objects.create(
            current_term=current_term,
            current_year=current_year,
            start_date=start_date,
            end_date=end_date,
            status=1  # Set as current term
        )
        messages.success(request, "New term settings have been added successfully")

        return redirect("settings")
    
    context = {
        'term_data' : Term.objects.filter(status=1)
    }
    return render(request, "frontend/settings.html", context)

# edit term
def edit_term(request):
    if request.method == "POST":
        id = request.POST.get("id")
        term = request.POST.get("term")
        # year = request.POST.get("year")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        this_term = Term.objects.get(id=id)

        this_term.current_term = term
        # this_term.current_year = year
        this_term.start_date = start_date
        this_term.end_date = end_date

        this_term.save()
        messages.success(request, "Academic term edit successfully")
        return redirect("settings")

# delete term
def delete_term(request):
    if request.method == "POST":
        id = request.POST.get("id")
        this_term = Term.objects.get(id=id)

        this_term.delete()
        messages.success(request, "Academic term deleted successfully")
        return redirect("settings")


# school information
@login_required
def school_info(request):
    if request.method == "POST":
        badge = request.FILES.get('badge')
        schoolname = request.POST.get("schoolname")
        contact = request.POST.get("contact")
        box_number = request.POST.get("box_number")
        email = request.POST.get("email")
        website = request.POST.get("website")

        # Check if there's existing school info data, assuming only one instance
        school_info = SchoolInfo.objects.first()

        if school_info is None:
            # If no data exists, create a new instance
            school_info = SchoolInfo(
                schoolname=schoolname,
                contact=contact,
                box_number=box_number,
                email=email,
                website=website,
            )
        else:
            # If data exists, update it
            messages.success(request, 'You can only add this information once!')
            return redirect("School Information")  

        if badge:
            fs = FileSystemStorage()
            image_filename = fs.save(badge.name, badge)
            school_info.badge = image_filename

        school_info.save()
        messages.success(request, 'School information updated successfully!')
        return redirect("School Information")  # Use the appropriate URL name for school_info page
    
    context = {
        'school_data': SchoolInfo.objects.all()
    }
    return render(request, "frontend/school_info.html", context)

def edit_school_info(request):
    if request.method == "POST":
        school_id = request.POST.get("id")
        school = get_object_or_404(SchoolInfo, pk=school_id)

        # Update school information based on form data
        school.schoolname = request.POST.get("schoolname")
        school.contact = request.POST.get("contact")
        school.box_number = request.POST.get("box_number")
        school.email = request.POST.get("email")
        school.website = request.POST.get("website")

        # Handle badge image update if necessary
        badge = request.FILES.get('badge')
        if badge:
            school.badge = badge

        school.save()

        # Add success message
        messages.success(request, 'School information updated successfully!')
        return redirect("School Information")

    # Render the school information template with context
    context = {
        'school_data': SchoolInfo.objects.all()
    }
    return render(request, "frontend/school_info.html", context)



def delete_teacher(request):
    if request.method == "POST":
        teacherid = request.POST.get("teacherid")

        teacher = Teachers.objects.get(pk=teacherid)
        teacher.delete()
        messages.success(request, "Teacher has been deleted")

        return redirect("Teachers List")

def edit_supportstaff(request):
    if request.method == 'POST':
        supportstaffid = request.POST.get("supportstaffid")
        supportstaffnames = request.POST.get('supportstaffnames')
        gender = request.POST.get('gender')
        contact = request.POST.get('contact')
        email = request.POST.get('email')
        address = request.POST.get('address')
        position = request.POST.get('position')
        qualification = request.POST.get('qualification')
        salary = request.POST.get('salary')
        bankaccnum = request.POST.get('bankaccnum')
        dob = request.POST.get('dob')
        joiningdate = request.POST.get('joiningdate')


        that_support_staff = Supportstaff.objects.get(pk=supportstaffid)

        # Update support staff attributes
        that_support_staff.supportstaffnames = supportstaffnames
        that_support_staff.gender = gender
        that_support_staff.contact = contact
        that_support_staff.email = email
        that_support_staff.address = address
        that_support_staff.position = position
        that_support_staff.qualification = qualification
        that_support_staff.salary = salary
        that_support_staff.bankaccnum = bankaccnum
        that_support_staff.dob = dob
        that_support_staff.joiningdate = joiningdate

        # Save the updated support staff
        that_support_staff.save()

        messages.success(request, "Support Staff edited successfully")
        return redirect("show supportstaff", supportstaffid=supportstaffid)
    # else:
    #     return render(request, 'your_template_name.html')  # Render a form for editing if the request method is GET

def delete_supportstaff(request):
    if request.method == "POST":
        supportstaffid = request.POST.get("supportstaffid")

        supportstaff = Supportstaff.objects.get(pk=supportstaffid)
        supportstaff.delete()
        messages.success(request, "Supportstaff has been deleted")

        return redirect("Support staff List")

# import students
from django.db.models import Q  # Import Q for complex queries

def import_students(request):
    if request.method == 'POST':
        students_file = request.FILES.get('student_file')

        if students_file:
            try:
                # Find the maximum stdnumber in the database
                max_stdnumber = Student.objects.aggregate(Max('stdnumber'))['stdnumber__max']

                # Generate the next stdnumber
                if max_stdnumber:
                    new_stdnumber = 'STD{:03d}'.format(int(max_stdnumber[3:]) + 1)
                else:
                    new_stdnumber = 'STD001'

                # Assuming the uploaded file is in Excel format (e.g., .xlsx)
                student_data = pd.read_excel(students_file)

                # Loop through the rows in the Excel file and create Student objects
                for _, row in student_data.iterrows():
                    # Check if a student with the same name and DOB already exists
                    existing_student = Student.objects.filter(
                        Q(childname=row["Child Name"]) &
                        Q(dob=row['Date of Birth'])
                    ).first()

                    if not existing_student:
                        class_name = row['Class']
                        this_class = Schoolclasses.objects.get(classname=class_name)

                        Student.objects.create(
                            stdnumber=new_stdnumber,
                            childname=row["Child Name"],
                            stdclass=this_class,
                            gender=row['Gender'],
                            dob=row['Date of Birth'],
                            address=row['Address'],
                            house=row['House'],
                            regdate=row['Registration Date'],
                            fathername=row["Father's Name"],
                            fcontact=row["Father's Contact"],
                            foccupation=row["Father's occupation"],
                            mothername=row["Mother's Name"],
                            mcontact=row["Mother's Contact"],
                            moccupation=row["Mother's Occupation"],
                            livingwith=row["Living with"],
                            guardianname=row["Guardian's Name"],
                            gcontact=row["Guardian's Contact"],
                        )

                        # Increment the new_stdnumber for the next student
                        new_stdnumber = 'STD{:03d}'.format(int(new_stdnumber[3:]) + 1)
                    else:
                        # Handle the case where the student already exists
                        messages.warning(
                            request,
                            f"Student {row['Child Name']} with DOB {row['Date of Birth']} already exists."
                        )

                # Optionally, you can add success messages or redirection
                messages.success(request, "Students imported successfully.")
                return redirect('AddStudents')  # Redirect to a page showing the student list
            except Exception as e:
                # Handle exceptions (e.g., file format error)
                messages.error(request, f"Error importing students: {str(e)}")
                return redirect('AddStudents')
        return redirect('AddStudents')

        

