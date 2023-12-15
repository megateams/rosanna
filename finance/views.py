from django.shortcuts import render, HttpResponse , redirect, get_object_or_404
from django.contrib import messages
from django.shortcuts import render, HttpResponse,redirect
from .models import *
from django.db.models import Sum
from django.db.models import Max
from frontend.models import *
from django.http import Http404
from django.core import serializers
from django.http import JsonResponse
from django.urls import reverse
# from django_excel_response import ExcelResponse
from django.db.models.functions import ExtractMonth
from django.db import transaction
import openpyxl
from datetime import date 
from datetime import datetime 
from collections import defaultdict
from django.contrib.auth import authenticate, login
from django.db.models import OuterRef, Subquery
from django.db.models import F
from frontend.views import encryptpassword

def financelogin(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        bursar = Administrators.objects.get(role = 'Bursar')

        if bursar.username == username and bursar.password == str(encryptpassword(password)):
            request.session['admin_id'] = bursar.id
            messages.success(request , "Login Successfull")
            return redirect("Finance Dashboard")
        else:
            messages.warning(request, 'Login Failed')
            return redirect('financeloginpage')
    return render(request , 'login.html')


def editteacherpayments(request):
    if request.method == 'POST':
        paymentid = request.POST.get('paymentid')
        teacherpayment = Teacherspayment.objects.get(paymentid=paymentid)
        salary = int(request.POST.get('salary'))
        amountpaid = int(request.POST.get('amountpaid'))
        teacherid = request.POST.get('teacherid')
        teachername = request.POST.get('teachername')

        # Calculate the new balance
        balance = salary - amountpaid

        # Fetch all payments made by the teacher
        payments = Teacherspayment.objects.filter(teacherid=teacherid)

        # Initialize accumulated payment with the edited amount paid
        new_accumulatedpayment = amountpaid

        # Calculate the new accumulated payment for all payments
        for payment in payments:
            if payment.paymentid == paymentid:
                # Skip the edited payment
                continue

            # Calculate the new accumulated payment for this payment
            new_accumulatedpayment += payment.amountpaid

        teacherpayment.salary = salary
        teacherpayment.amountpaid = amountpaid
        teacherpayment.balance = balance
        teacherpayment.teacherid = teacherid
        teacherpayment.teachername = teachername

        # Update the accumulated payment
        teacherpayment.accumulatedpayment = new_accumulatedpayment
        teacherpayment.save()

        messages.success(request, "Teacher Payment Edit Successful")
        return redirect('teacherpaymentslists')

    # Rest of your code for GET requests



def editsupportstaffpayment(request):
    if request.method == 'POST':
        paymentid = request.POST.get('paymentid')
        staffpayment = Supportstaffpayment.objects.get(paymentid = paymentid)
        supportstaffid = request.POST.get('supportstaffid')
        salary = request.POST.get('salary')
        amountpaid = request.POST.get('amountpaid')
        #balance = request.POST.get('balance')
        #paymentmethod = request.POST.get('paymentmethod')
        #bankaccnum = request.POST.get('bankaccnum')
        paymentdate = request.POST.get('paymentdate')
        staffname = request.POST.get('staffname')

        staffpayment.supportstaffid = supportstaffid
        staffpayment.salary = salary
        staffpayment.amountpaid = amountpaid 
        staffpayment.balance = int(salary) - int(amountpaid)
        #staffpayment.paymentmethod = paymentmethod
        #staffpayment.bankaccnum = bankaccnum
        staffpayment.paymentdate = paymentdate 
        staffpayment.staffname = staffname 

        staffpayment.save

        messages.success(request , 'Edited Successfully')
        return redirect('SupportstaffpaymentsLists')
    return render(request , 'finance/staffpayments/editsupportstaffpayments.html' , {'staffpayment': staffpayment})

def deleteteacherpayment(request):
    if request.method == 'POST':
        paymentid = request.POST.get('paymentid')
        payid = Teacherspayment.objects.filter(paymentid = paymentid)
        payid.delete()
        messages.success(request , 'Payment Deleted Successfully')
        return redirect('teacherpaymentslists')

def deletesupportstaffpayment(request):
    if request.method == 'POST':
        paymentid = request.POST.get('paymentid')
        payid = Supportstaffpayment.objects.filter(paymentid = paymentid)
        payid.delete()
        messages.success(request , "Payment deleted successfully")
        supportstaffinfo = Supportstaffpayment.objects.all()
        return redirect('SupportstaffpaymentsLists')

def financeaddStaffpayments(request):
    if request.method == 'POST':
        supportstaffid = request.POST.get('support-staffid')
        supportstaffname = request.POST.get('support-staffname')
        paymentdate = request.POST.get('datepaid')
        salary = request.POST.get('salary')
        amount = request.POST.get('amount')

        Supportstaffpayment.objects.create(
            supportstaffid = supportstaffid ,
            salary = salary ,
            amountpaid = amount ,
            #balance = balance ,
            supportstaffname = supportstaffname,
            #paymentdate = datepaid,
        )
        Supportstaffpayment.save()
    return render(request , 'finance/staffpayments/financesupportstaffpaymentsList.html')

def receipts(request):
    if request.method == 'POST':
        receiptnum = request.POST.get('receiptnum')
        transactiondate = request.POST.get('transactiondate')
        amountpaid = request.POST.get('amountpaid')
        item = request.POST.get('item')
        balance = request.POST.get('balance')
        payername = request.POST.get('payername')

        Receipts.objects.create(
            receiptnum = receiptnum ,
            transactiondate = transactiondate ,
            amountpaid = amountpaid ,
            item = item ,
            balance = balance ,
            payername = payername ,
        )

        Receipts.save()
    return render(request, 'finace/financedashboard.html')

def bankdetails(request):
    if request.method == 'POST':
        staffname = request.POST.get('staffname')
        bankname = request.POST.get('bankname')
        accnum = request.POST.get('accnum')
        accname = request.POST.get('accname')

        Bankdetails.objects.create(
            staffname = staffname ,
            bankname = bankname ,
            accnum = accnum ,
            accname = accname
        )

        Bankdetails.save()
    return render(request , 'finance/financedashboard.html')

def staffpayments(request):
    if request.method == 'POST':
        staffname = request.POST.get('staffname')
        datepaid = request.POST.get('datepaid')
        salary = request.POST.get('salary')
        amountpaid = request.POST,get('amountpaid')
        balance = request.POST.get('balance')
        position = request.POST.get('position')
        bankaccnum = request.POST.get('bankaccnum')
        
        Staffpayments.objects.create(
            staffname = staffname ,
            datepaid = datepaid ,
            salary = salary ,
            amountpaid = amountpaid ,
            balance = balance ,
            position = position ,
            bankaccnum = bankaccnum
        )
        
        Staffpayments.save()
    return render(request , 'finace/financedashboard.html')
        



def students_list(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)
    # Retrieve a list of students from your database
    studentslist = Student.objects.all() 
    classes = Schoolclasses.objects.all() 

    # Pass the list of students to the template for rendering
    context = {
        'studentslist': studentslist,
        'classes': classes,
        'bursar': bursar,
        }
    return render(request, 'finance/students.html', context)

def students_by_class(request, class_id):
    # Retrieve all available classes
    classes = Schoolclasses.objects.all()

    try:
        # Fetch the selected class based on the class_id
        selected_class = Schoolclasses.objects.get(classid=class_id)

        # Fetch students in the selected class
        students_in_class = Student.objects.filter(stdclass=selected_class.classid)
    except Schoolclasses.DoesNotExist:
        selected_class = None
        students_in_class = []

    return render(request, 'finance/studentsbyclass.html', {
        'studentslist': students_in_class,
        'classes': classes,
        'selected_class': selected_class,
    })

def assign_school_code(request, stdnumber):
    student = get_object_or_404(Student, stdnumber=stdnumber)
    error_message = None
    
    if request.method == 'POST':
        school_code = request.POST.get('schoolpaycode')
        if school_code:
            # Check if the school pay code is already assigned to another student
            existing_student = Student.objects.filter(schoolpaycode=school_code).exclude(stdnumber=stdnumber).first()
            if existing_student:
                # If the code is assigned to another student, set the error message
                messages.success(request,"The School Pay Code  is already assigned to another student.")
            else:
                # Assign the code to the current student
                student.schoolpaycode = school_code
                student.save()

            return redirect("students_list")

# Create your views here.
def financedashboard(request):
    # Check if the bursar is authenticated (if you are using sessions)
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view



    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    term_data = Term.objects.get(status=1)
    terms = Term.objects.all()

    bursar = Administrators.objects.get(id=admin_id)
    utilities = Utilities.objects.filter(term=term_data.current_term, year=term_data.current_year)
    total_amount_paid = utilities.aggregate(Sum('amountpaid'))['amountpaid__sum']
    if total_amount_paid == None:
        total_amount_paid = 0

    supplementary = Supplementaryincome.objects.filter(term=term_data.current_term, year=term_data.current_year)
    total_amt = supplementary.aggregate(Sum('amount'))['amount__sum']
    if total_amt == None:
        total_amt = 0
    

    fees = Fees.objects.filter(term=term_data.current_term, year=term_data.current_year)
    total_amount = fees.aggregate(Sum('amount'))['amount__sum']
    if total_amount == None:
        total_amount = 0

    # Retrieve only the latest balance for each student based on the latest timestamp and date

    latest_dates = Fees.objects.filter(stdnumber=OuterRef('stdnumber')).order_by('-date')
    fees_list = Fees.objects.annotate(
        latest_date=Subquery(latest_dates.values('date')[:1])
    ).filter(balance__gt=0.0, date=F('latest_date'),term=term_data.current_term, year=term_data.current_year)

    
    sspayments = Supportstaffpayment.objects.filter(term=term_data.current_term, year=term_data.current_year)
    total_sspayments = sspayments.aggregate(Sum('amountpaid'))['amountpaid__sum']
    if total_sspayments == None:
        total_sspayments = 0

    trpayments = Teacherspayment.objects.filter(term=term_data.current_term, year=term_data.current_year)
    total_trpayments = trpayments.aggregate(Sum('amountpaid'))['amountpaid__sum']
    if total_trpayments == None:
        total_trpayments = 0

    # Calculate total income (fees)
    total_income = total_amount + total_amt

    # Calculate total expenses (teacher payments + support staff payments + utilities)
    total_expenses = total_trpayments + total_sspayments + total_amount_paid

    # Calculate profit
    profit = total_income - total_expenses
    # total = total_amount + total_amount_paid + total_sspayments + total_trpayments
    # if total == 0:
    #     fees_percentage = 0
    #     utilities_percentage = 0
    #     sspayments_percentage = 0
    #     trpayments_percentage = 0
    # else:


    # Calculate the percentages
    if total_amount == 0 or total_amount_paid== 0 or total_sspayments==0 or total_trpayments==0: 
        context = {
            'total_amount_paid': total_amount_paid,
            'total_amount': total_amount,
            'total_sspayments' : total_sspayments,
            'total_trpayments' : total_trpayments,
            'term_data' : term_data,
            'fees_list': fees_list,
            'bursar' : bursar,
        }
        return render(request, "finance/financedashboard.html", context)
    else: 
        total = total_amount + total_amount_paid + total_sspayments + total_trpayments

    total = total_amount + total_amount_paid + total_sspayments + total_trpayments
    if total == 0:
        fees_percentage = 0
        utilities_percentage = 0
        sspayments_percentage = 0
        trpayments_percentage = 0
    else:
        fees_percentage = (total_amount / total) * 100
        utilities_percentage = (total_amount_paid / total) * 100
        sspayments_percentage = (total_sspayments / total) * 100
        trpayments_percentage = (total_trpayments / total) * 100

    context = {
        'total_amount_paid': total_amount_paid,
        'total_amount': total_amount,
        'total_amt': total_amt,
        'total_sspayments' : total_sspayments,
        'total_trpayments' : total_trpayments,
        'term_data' : term_data,
        'fees_list': fees_list,
        'fees_percentage': fees_percentage,
        'utilities_percentage': utilities_percentage,
        'sspayments_percentage': sspayments_percentage,
        'trpayments_percentage': trpayments_percentage,
        'bursar' : bursar,
        'total_income' : total_income,
        'total_expenses' : total_expenses,
        'profit' : profit,
        'terms' : terms,
    }
    return render(request, "finance/financedashboard.html", context)


# that term
def that_term(request, id):
    # Check if the bursar is authenticated (if you are using sessions)
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view



    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    term_data = Term.objects.get(id=id)
    terms = Term.objects.all()

    bursar = Administrators.objects.get(id=admin_id)
    utilities = Utilities.objects.filter(term=term_data.current_term, year=term_data.current_year)
    total_amount_paid = utilities.aggregate(Sum('amountpaid'))['amountpaid__sum']
    if total_amount_paid == None:
        total_amount_paid = 0

    supplementary = Supplementaryincome.objects.filter(term=term_data.current_term, year=term_data.current_year)
    total_amt = supplementary.aggregate(Sum('amount'))['amount__sum']
    if total_amt == None:
        total_amt = 0

    fees = Fees.objects.filter(term=term_data.current_term, year=term_data.current_year)
    total_amount = fees.aggregate(Sum('amount'))['amount__sum']
    if total_amount == None:
        total_amount = 0

    # Retrieve only the latest balance for each student based on the latest timestamp and date

    latest_dates = Fees.objects.filter(stdnumber=OuterRef('stdnumber')).order_by('-date')
    fees_list = Fees.objects.annotate(
        latest_date=Subquery(latest_dates.values('date')[:1])
    ).filter(balance__gt=0.0, date=F('latest_date'),term=term_data.current_term, year=term_data.current_year)

    
    sspayments = Supportstaffpayment.objects.filter(term=term_data.current_term, year=term_data.current_year)
    total_sspayments = sspayments.aggregate(Sum('amountpaid'))['amountpaid__sum']
    if total_sspayments == None:
        total_sspayments = 0

    trpayments = Teacherspayment.objects.filter(term=term_data.current_term, year=term_data.current_year)
    total_trpayments = trpayments.aggregate(Sum('amountpaid'))['amountpaid__sum']
    if total_trpayments == None:
        total_trpayments = 0

    # Calculate total income (fees)
    total_income = total_amount

    # Calculate total expenses (teacher payments + support staff payments + utilities)
    total_expenses = total_trpayments + total_sspayments + total_amount_paid

    # Calculate profit
    profit = total_income - total_expenses
    total = total_amount + total_amount_paid + total_sspayments + total_trpayments
    if total == 0:
        fees_percentage = 0
        utilities_percentage = 0
        sspayments_percentage = 0
        trpayments_percentage = 0
    else:
        fees_percentage = (total_amount / total) * 100
        utilities_percentage = (total_amount_paid / total) * 100
        sspayments_percentage = (total_sspayments / total) * 100
        trpayments_percentage = (total_trpayments / total) * 100

    context = {
        'total_amount_paid': total_amount_paid,
        'total_amount': total_amount,
        'total_amt': total_amt,
        'total_sspayments' : total_sspayments,
        'total_trpayments' : total_trpayments,
        'term_data' : term_data,
        'fees_list': fees_list,
        'fees_percentage': fees_percentage,
        'utilities_percentage': utilities_percentage,
        'sspayments_percentage': sspayments_percentage,
        'trpayments_percentage': trpayments_percentage,
        'bursar' : bursar,
        'total_income' : total_income,
        'total_expenses' : total_expenses,
        'profit' : profit,
        'terms' : terms,
    }
    return render(request, "finance/financedashboard.html", context)

# @login_required

def bursar_profile(request):
    # Check if the user is authenticated
    if request.user.is_authenticated:
        # Query the Administrators model to get the Bursar administrator
        try:
            bursar = Administrators.objects.get(role='Bursar')
        except Administrators.DoesNotExist:
            # Handle the case where there is no Bursar in the database
            raise Http404("Bursar profile not found")
        
        context = {
            'bursar': bursar,
        }

        return render(request, 'finance/profile.html', context)
    else:
        # Handle the case where the user is not authenticated
        # You can redirect or show an error message here
        pass

def edit_bursar_profile(request):
    try:
        # Retrieve the Bursar's data based on their role
        bursar = Administrators.objects.get(role='Bursar')
    except Administrators.DoesNotExist:
        raise Http404("Bursar profile not found")

    if request.method == 'POST':
        username = request.POST.get("username")
        new_pass = request.POST.get("new_password")
        confirm_pass = request.POST.get("confirm_password")

        if new_pass == confirm_pass:
            # Update the Bursar's password (replace this with your actual password update logic)
            bursar.password = encryptpassword(new_pass)
            bursar.username = username
            bursar.save()

            messages.success(request, "Profile updated successfully")
        else:
            messages.error(request, "Passwords do not match")

        return redirect("bursar_profile")

    return render(request, 'finance/profile.html', {'bursar': bursar})



# fees views
def financeaddFees(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)

    if request.method == 'POST':
        stdnumber = request.POST.get('stdnumber')
        stdname = Student.objects.get(stdnumber=stdnumber).childname  # Get student name
        studentclass = request.POST.get('studentclass')
        classfees = request.POST.get("classfees")
        amount = request.POST.get("amount")  # Get amount from fees structure
        modeofpayment = request.POST.get('modeofpayment')
        # date = request.POST.get('date')
        # timestamp = request.POST.get('timestamp')
        
        # Capture the current date and time
        current_date = date.today()
        current_time = datetime.now().time()
        print(current_time)

        term_data = Term.objects.get(status=1)
        # Calculate balance
        balance = int(classfees) - int(amount)

        if int(amount) <= int(classfees):
            # Check if there are previous fee records for the student
            last_fee = Fees.objects.filter(stdnumber_id=stdnumber).last()

            if last_fee is None:
                # No previous fee records exist, create a new entry
                fees = Fees.objects.create(
                    stdnumber_id=stdnumber,
                    stdname=stdname,
                    studentclass=studentclass,
                    classfees=classfees,
                    amount=amount,
                    balance=balance,
                    modeofpayment=modeofpayment,
                    accumulatedpayment=amount,
                    term = term_data.current_term,
                    year = term_data.current_year
                )
                fees.save()
            else:
                if last_fee.balance == 0:
                    # Previous balance was 0, create a new entry
                    fees = Fees.objects.create(
                        stdnumber_id=stdnumber,
                        stdname=stdname,
                        studentclass=studentclass,
                        classfees=classfees,
                        amount=amount,
                        balance=balance,
                        modeofpayment=modeofpayment,
                        accumulatedpayment=amount,
                        term = term_data.current_term,
                        year = term_data.current_year
                    )
                    fees.save()
                else:
                    # Previous balance was greater than 0, update accumulated payment and balance
                    accumulatedpayment = last_fee.accumulatedpayment + int(amount)
                    new_balance = int(classfees) - accumulatedpayment
                    if new_balance < 0:
                        messages.error(request, 'Student has cleared all fees for this term')
                    else:
                        fees = Fees.objects.create(
                            stdnumber_id=stdnumber,
                            stdname=stdname,
                            studentclass=studentclass,
                            classfees=classfees,
                            amount=amount,
                            balance=new_balance,
                            modeofpayment=modeofpayment,
                            accumulatedpayment=accumulatedpayment,
                            term = term_data.current_term,
                            year = term_data.current_year
                        )
                        fees.save()

            messages.success(request, 'Fees registered successfully.')
            return redirect('Add Fees')
        else:
            messages.error(request, 'Amount is greater than class fees. Please try again')

    students = Student.objects.all()
    fees_structures = Feesstructure.objects.all()
    return render(request, 'finance/fees/financeaddFees.html', {'students': students, 'fees_structures': fees_structures, 'bursar': bursar})

def financefeesList(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    term_data = Term.objects.get(status=1)
    terms = Term.objects.all()
    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)

    fees_list = Fees.objects.filter(term=term_data.current_term, year=term_data.current_year)
    total_amount = fees_list.aggregate(Sum('amount'))['amount__sum']
    classes = Schoolclasses.objects.all()
    context = {
        'fees_list': fees_list,
        'total_amount': total_amount,
        'classes': classes,
        'bursar': bursar,
        'terms': terms
    }
    return render(request,'finance/fees/financefeesList.html',context)

def fees_by_class(request, class_id):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)

    classes = Schoolclasses.objects.all()
    term_data = Term.objects.get(status=1)
    
    try:
        # Fetch the selected class based on the class_id
        selected_class = Schoolclasses.objects.get(classid=class_id)
        
        # Fetch fees records for students in the selected class
        fees_list = Fees.objects.filter(studentclass=selected_class.classname, term=term_data.current_term, year=term_data.current_year)
        
        # Retrieve the class fees from the Fees model
        total_amount = fees_list.aggregate(Sum('amount'))['amount__sum']
        # classfees = classfees.amount
    except Schoolclasses.DoesNotExist:
        selected_class = None
        fees_list = []
        classfees = 0  # Default value if class is not found or fees not set
    
    return render(request, 'finance/fees/fees_by_class.html', {
        'fees_list': fees_list,
        'classes': classes,
        'selected_class': selected_class,
        'total_amount': total_amount,  # Pass the class fees to the template
        'bursar' : bursar
    })

def delete_fee(request):
    if request.method == 'POST':
        paymentid = request.POST.get("paymentid")
        fee = Fees.objects.get(paymentid=paymentid)
        fee.delete()
        messages.success(request, f"Fee record {paymentid} has been deleted.")
        return redirect('Fees List')  # Adjust this to the correct URL name

    return redirect('Fees List')  # Adjust this to the correct URL name

def edit_std_fees(request):
    if request.method == 'POST':
        paymentid = request.POST.get("paymentid")
        amount = float(request.POST.get("amount"))
        modeofpayment = request.POST.get("modeofpayment")
        fee = Fees.objects.get(paymentid=paymentid)

        # Fetch the existing balance and accumulated payment from the database and convert them to float
        balance = float(fee.balance)
        accumulatedpayment = float(fee.accumulatedpayment)

        # Explicitly convert fee.classfees to float
        classfees = float(fee.classfees)

        # Calculate the new accumulated payment
        new_accumulatedpayment = accumulatedpayment - float(fee.amount) + amount

        # Calculate the new balance based on the class fees and the new accumulated payment
        balance = classfees - new_accumulatedpayment

        # Update the fee record
        fee.amount = amount
        fee.balance = balance
        fee.modeofpayment = modeofpayment
        fee.accumulatedpayment = new_accumulatedpayment
        fee.save()

        messages.success(request, f"Fee record {paymentid} has been edited.")
        return redirect('Fees List')  # Adjust this to the correct URL name

    # Rest of your code for GET requests



def feesclearedstudents(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)

    
    term_data = Term.objects.get(status=1)
    # Filter the Fees model to get students with a balance of 0
    cleared_students = Fees.objects.filter(balance=0 ,term=term_data.current_term, year=term_data.current_year)
    classes = Schoolclasses.objects.all()

    # Pass the cleared_students queryset to the template
    return render(request, 'finance/fees/feesclearedstudents.html', {'cleared_students': cleared_students, 'classes': classes, 'bursar': bursar})

def feesclearedstudents_byclass(request, class_id):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)
    classes = Schoolclasses.objects.all()
    selected_class = None  # Initialize selected_class

    try:
        # Fetch the selected class based on the class_id
        selected_class = Schoolclasses.objects.get(classid=class_id)

        # Fetch cleared students for the selected class (balance = 0)
        cleared_students = Fees.objects.filter(studentclass=selected_class.classname, balance=0)
    except Schoolclasses.DoesNotExist:
        pass  # Handle the case where the class is not found

    return render(request, 'finance/fees/feesclearedstudents_byclass.html', {
        'cleared_students': cleared_students,
        'classes': classes,
        'selected_class': selected_class,
        'bursar':bursar
    })

def generate_clearance(request, stdnumber):
    try:
        # Get the student's information based on their student number
        student = get_object_or_404(Fees, stdnumber=stdnumber)

        # Perform clearance generation logic here
        # For example, you can update the student's clearance status
        student.clearance_status = True  # Assuming you have a field for clearance status in the Fees model
        student.save()

        # Redirect to the clearance card view with the student's stdnumber
        return redirect('Clearance Card', stdnumber=stdnumber)

        # Optionally, you can add a success message
        messages.success(request, f"Clearance generated for {student.stdname}.")

    except Fees.DoesNotExist:
        # Handle the case where the student is not found
        messages.error(request, "Student not found.")

    # Redirect back to the cleared students list or any other appropriate page
    return redirect('Cleared Students List')  # Adjust this URL name as needed

def clearance_card(request, stdnumber):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)
    # Retrieve the student's information based on the stdnumber
    student = Fees.objects.get(stdnumber=stdnumber,balance=0)
    
    # get more student details
    student_img = Student.objects.get(stdnumber=stdnumber)

    # get school information
    term_data = Term.objects.get(status=1)
    # Render the clearance card template and pass the student's information

    context={
        'student': student,
        'student_img': student_img,
        'term_data':term_data,
        'bursar' : bursar
    }
    return render(request, 'finance/fees/clearancecard.html',context)

# fees views

def financeaddFeesstructure(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)
    if request.method == 'POST':
        classname = request.POST.get('classname')
        amount = request.POST.get('amount')
        term_data = Term.objects.get(status=1)

        term = term_data.current_term
        year = term_data.current_year


        this_class = Feesstructure.objects.filter(classname = classname)
        if(this_class):
            messages.success(request, "This class already exists")
        else:
            fees_structure = Feesstructure.objects.create(
                classname = classname, 
                amount = amount,
                term = term,
                year = year
                )
            fees_structure.save()
            messages.success(request, f"Fees Structure for class '{classname}' added successfully.")
        return redirect('Add Fees Structure')
    classes = Schoolclasses.objects.all()
    return render(request, 'finance/feesstructure/financeaddFeesstructure.html',{"classes":classes,'bursar': bursar})

def financefeesstructureList(request):
    if 'admin_id' not in request.session:
        
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)
    fees_list = Feesstructure.objects.all()
    context = {
        'fees_list': fees_list,
        'bursar' : bursar,
    }

    return render(request, 'finance/feesstructure/financefeesstructureList.html', context)

def deletefeesstructure(request, feesstructureid):
    try:
        fees_structure = Feesstructure.objects.get(pk=feesstructureid)
        fees_structure.delete()
        messages.success(request, 'Fees Structure deleted successfully.')
    except Feesstructure.DoesNotExist:
        messages.error(request, 'Fees Structure not found.')
    return redirect('Fees Structure List')

def editfeesstructure(request, feesstructureid):
    try:
        fees_structure = Feesstructure.objects.get(pk=feesstructureid)
        if request.method == 'POST':
            updated_classname = request.POST.get('classname')
            updated_amount = request.POST.get('amount')
            fees_structure.classname = updated_classname
            fees_structure.amount = updated_amount
            fees_structure.save()
            messages.success(request, 'Fees Structure updated successfully.')
            return redirect('Fees Structure List')
        context = {'fees_structure': fees_structure}
        return render(request, 'finance/feesstructure/editfeesstructure.html', context)
    except Feesstructure.DoesNotExist:
        messages.error(request, 'Fees Structure not found.')
        return redirect('Fees Structure List')

# teacherpayments views
def financeaddTeacherpayments(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)
    teachersdata = Teachers.objects.all()

    if request.method == 'POST':
        teacherid = request.POST.get('teacherid')
        term_data = Term.objects.get(status=1)

        # Capture the current date
        current_date = date.today()

        teacher = Teachers.objects.get(teacherid=teacherid)
        teachername = teacher.teachernames
        # paymentdate = request.POST.get('paymentdate')
        salary = float(teacher.salary)
        amountpaid = float(request.POST.get('amount'))
        term = term_data.current_term
        year = term_data.current_year
        
        if amountpaid > float(teacher.salary):
            messages.error(request , 'Payment not Added')
            return render(request, 'finance/staffpayments/financeaddTeacherpayments.html', {'balancealert':"Amount Paid must not be greater than the Salary" , 'teachers': teachersdata})

        payment = Teacherspayment.objects.filter(teacherid = teacherid).last()

        print(payment.paymentdate)

        if payment == None:
            Teacherspayment.objects.create(
                teacherid = teacherid,
                teachername = teachername,
                paymentdate = current_date,
                salary = salary,
                amountpaid = amountpaid,
                accumulatedpayment = amountpaid,
                balance = float(teacher.salary) - amountpaid,
                term = term,
                year = year
            )

            messages.success(request, 'Teacher payment added successfully.')
            return render(request, 'finance/staffpayments/financeaddTeacherpayments.html', {'teachers': teachersdata, 'bursar': bursar})

        if payment.balance == 0:
            
            if payment.balance == 0 and payment.paymentdate.month == current_date.month and payment.paymentdate.year == current_date.year:
                messages.success(request, 'Teacher already paid for this month.')

            else:
                Teacherspayment.objects.create(
                    teacherid = teacherid,
                    teachername = teachername,
                    paymentdate = current_date,
                    salary = salary,
                    amountpaid = amountpaid,
                    accumulatedpayment = amountpaid,
                    balance = float(teacher.salary) - amountpaid,
                    term = term,
                    year = year
                )
                messages.success(request, 'Payment successfull.')
            return render(request, 'finance/staffpayments/financeaddTeacherpayments.html', {'teachers': teachersdata, 'bursar': bursar})


            

            

        if payment.balance > 0:
            accumulatedpayment = payment.accumulatedpayment + amountpaid
            newbalance = payment.salary - accumulatedpayment
            if newbalance < 0:
                messages.error(request , 'Payment not Added')
                return render(request, 'finance/staffpayments/financeaddTeacherpayments.html', {'balancealert':"Amount Paid must not be greater than the Salary" , 'teachers': teachersdata})
            else:
                Teacherspayment.objects.create(
                    teacherid = teacherid,
                    teachername = teachername,
                    paymentdate = current_date,
                    salary = salary,
                    amountpaid = amountpaid,
                    accumulatedpayment = accumulatedpayment,
                    balance = newbalance,
                    term = term,
                    year = year
                )
                messages.success(request, 'Teacher payment added successfully.')
                return render(request, 'finance/staffpayments/financeaddTeacherpayments.html', {'teachers': teachersdata,'bursar': bursar})

    return render(request, 'finance/staffpayments/financeaddTeacherpayments.html', {'teachers': teachersdata,'bursar':bursar})

def financeteacherpaymentsList(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']
    term_data = Term.objects.get(status=1)

    bursar = Administrators.objects.get(id=admin_id)
    total_trpayments = Teacherspayment.objects.filter(term=term_data.current_term, year=term_data.current_year).aggregate(Sum('amountpaid'))['amountpaid__sum']
    teacherspayment = Teacherspayment.objects.filter(term=term_data.current_term, year=term_data.current_year)
    context = {
       'teachers':teacherspayment,
       'total_trpayments': total_trpayments, 
       'bursar': bursar
    }
    return render(request,'finance/staffpayments/financeteacherpaymentsList.html' ,context)
# teacherpayments views

# supportstaffpayments views
def financeaddsupportstaffpayments(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)
    supportstaffdata = Supportstaff.objects.all()
    
    if request.method == 'POST':
        term_data = Term.objects.get(status=1)
        supportstaffid = request.POST.get('support-staffid')
        supportstaffname = Supportstaff.objects.get(supportstaffid = supportstaffid)
        staffnames = supportstaffname.supportstaffnames
        # paymentdate = request.POST.get('paymentdate')
        salary = float(request.POST.get('salary'))
        amountpaid = float(request.POST.get('amountpaid'))
        term = term_data.current_term
        year = term_data.current_year

        # Capture the current date
        current_date = date.today()


        if amountpaid > float(supportstaffname.salary):
            messages.error(request , 'Payment not Added')
            return render(request, 'finance/staffpayments/financeaddsupportstaffpayments.html', {'balancealert':"Amount Paid must not be greater than the Salary" , 'supportstaff': supportstaffdata, 'bursar': bursar})

        payment = Supportstaffpayment.objects.filter(supportstaffid = supportstaffid).last()

        if payment == None:
            Supportstaffpayment.objects.create(
                supportstaffid = supportstaffid ,
                staffname = staffnames ,
                paymentdate = current_date,
                salary = salary,
                amountpaid = amountpaid,
                accumulatedamount = amountpaid,
                balance = float(supportstaffname.salary) - amountpaid,
                term = term,
                year = year
            )

            messages.success(request, 'Support Staff payment added successfully.')
            return render(request, 'finance/staffpayments/financeaddsupportstaffpayments.html', {'supportstaff': supportstaffdata, 'bursar' :bursar})
        
        if payment.balance == 0:
            Supportstaffpayment.objects.create(
                supportstaffid = supportstaffid ,
                staffname = staffnames ,
                paymentdate = current_date,
                salary = salary,
                amountpaid = amountpaid,
                accumulatedamount = amountpaid,
                balance = float(supportstaffname.salary) - amountpaid,
                term = term,
                year = year
            )

            messages.success(request, 'Support staff payment added successfully.')
            return render(request, 'finance/staffpayments/financeaddsupportstaffpayments.html', {'supportstaff': supportstaffdata, 'bursar':bursar})
        
        if payment.balance > 0:
            accumulatedpayment = payment.accumulatedamount + amountpaid
            newbalance = payment.salary - accumulatedpayment
            if newbalance < 0:
                messages.error(request , 'Payment not Added')
                return render(request, 'finance/staffpayments/financeaddsupportstaffpayments.html', {'balancealert':"Amount Paid must not be greater than the Salary" , 'supportstaff': supportstaffdata})
            else:
                Supportstaffpayment.objects.create(
                    supportstaffid = supportstaffid ,
                    staffname = staffnames ,
                    paymentdate = current_date,
                    salary = salary,
                    amountpaid = amountpaid,
                    accumulatedamount = accumulatedpayment,
                    balance = newbalance,
                    term = term,
                year = year
                )
                messages.success(request, 'Teacher payment added successfully.')
                return render(request, 'finance/staffpayments/financeaddsupportstaffpayments.html', {'teachers': supportstaffdata, 'bursar':bursar})
    return render(request, 'finance/staffpayments/financeaddsupportstaffpayments.html', {'supportstaff': supportstaffdata, 'bursar': bursar})

def financesupportstaffpaymentsList(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']
    term_data = Term.objects.get(status=1)

    bursar = Administrators.objects.get(id=admin_id)
    total_sspayments = Supportstaffpayment.objects.filter(term=term_data.current_term, year=term_data.current_year).aggregate(Sum('amountpaid'))['amountpaid__sum']
    support_staff_payments = Supportstaffpayment.objects.filter(term=term_data.current_term, year=term_data.current_year)
    context = {
        'supportstaffpayments': support_staff_payments,
        'total_sspayments': total_sspayments,
        'bursar': bursar
    }

    return render(request, 'finance/staffpayments/financesupportstaffpaymentsList.html', context)

# supportstaffpayments views

# utilities views
def financeaddUtilities(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)
    if request.method == 'POST':
        term_data = Term.objects.get(status=1)

        # Capture the current date
        current_date = date.today()

        term = term_data.current_term
        year = term_data.current_year
        category = request.POST.get('category')
        amountpaid = request.POST.get('amountpaid')
        utilities = Utilities(
            category=category,
            utilitiesdate=current_date,
            amountpaid=amountpaid,
            term = term,
            year = year
        )
        utilities.save()
        messages.success(request, 'Utility added successfully.')  # Display a success message
        return redirect('Add Utilities')  # Redirect to utilities list page

    return render(request, 'finance/utilities/financeaddUtilities.html', {'bursar':bursar})

def financeutilitiesList(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']
    term_data = Term.objects.get(status=1)

    bursar = Administrators.objects.get(id=admin_id)
    total_amount_paid = Utilities.objects.filter(term=term_data.current_term, year=term_data.current_year).aggregate(Sum('amountpaid'))['amountpaid__sum']
    utilities = Utilities.objects.filter(term=term_data.current_term, year=term_data.current_year)
    context = {
        'utilities': utilities,
        'total_amount_paid': total_amount_paid,
        'bursar': bursar
    }
    return render(request, 'finance/utilities/financeutilitiesList.html', context)

def delete_utilities(request):
    if request.method == 'POST':
        utilitiesid = request.POST.get("utilitiesid")
        try:
            utility = Utilities.objects.get(utilitiesid=utilitiesid)
            utility.delete()
            messages.success(request, f"Utility record {utilitiesid} has been deleted.")
        except Utilities.DoesNotExist:
            messages.error(request, f"Utility record {utilitiesid} does not exist.")

    return redirect('Utilities List')  # Adjust this to the correct URL name


def edit_utilities(request, utilitiesid):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)
    try:
        utility = Utilities.objects.get(utilitiesid=utilitiesid)

        if request.method == 'POST':
            updated_category = request.POST.get('category')
            updated_utilitiesdate = request.POST.get('utilitiesdate')
            updated_amountpaid = request.POST.get('amountpaid')
            utility.category = updated_category
            utility.amountpaid = updated_amountpaid
            utility.save()
            messages.success(request, 'Utility updated successfully.')
            return redirect('Add Utilities')

        context = {'utility': utility, 'bursar': bursar}
        return render(request, 'finance/utilities/edit_utilities.html', context)

    except Utilities.DoesNotExist:
        messages.error(request, 'Utility not found.')
        return redirect('Utilities List')
# utilities views

# Supplementary Income views
def addSupplementaryincome(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)
    if request.method == 'POST':
        term_data = Term.objects.get(status=1)

        # Capture the current date
        current_date = date.today()

        term = term_data.current_term
        year = term_data.current_year
        category = request.POST.get('category')
        amount = request.POST.get('amount')
        supplementary = Supplementaryincome(
            category=category,
            date=current_date,
            amount=amount,
            term = term,
            year = year
        )
        supplementary.save()
        messages.success(request, 'Supplementary income added successfully.')  # Display a success message
        return redirect('Add Supplementary Income')  # Redirect to utilities list page

    return render(request, 'finance/supplementaryincome/addSupplementaryincome.html', {'bursar':bursar})

def supplementaryincomeList(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']
    term_data = Term.objects.get(status=1)

    bursar = Administrators.objects.get(id=admin_id)
    total_amt = Supplementaryincome.objects.filter(term=term_data.current_term, year=term_data.current_year).aggregate(Sum('amount'))['amount__sum']
    supplementary = Supplementaryincome.objects.filter(term=term_data.current_term, year=term_data.current_year)
    context = {
        'supplementaryincomes': supplementary,
        'total_amt': total_amt,
        'bursar': bursar
    }
    return render(request, 'finance/supplementaryincome/supplementaryincomeList.html', context)

def delete_supplementary(request):
    if request.method == 'POST':
        supplementaryincomeid = request.POST.get("supplementaryincomeid")
        try:
            supplementary = Supplementaryincome.objects.get(supplementaryincomeid=supplementaryincomeid)
            supplementary.delete()
            messages.success(request, f"Supplementaryincome record {supplementaryincomeid} has been deleted.")
        except Supplementaryincome.DoesNotExist:
            messages.error(request, f"Supplementaryincome record {supplementaryincomeid} does not exist.")

    return redirect('Supplementary Income List')  # Adjust this to the correct URL name


def edit_supplementary(request, supplementaryincomeid):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)
    try:
        supplementary = Supplementaryincome.objects.get(supplementaryincomeid=supplementaryincomeid)

        if request.method == 'POST':
            updated_category = request.POST.get('category')
            updated_date = request.POST.get('date')
            updated_amount = request.POST.get('amount')
            supplementary.category = updated_category
            supplementary.amount = updated_amount
            supplementary.save()
            messages.success(request, 'Supplementary updated successfully.')
            return redirect('Add supplementary')

        context = {'supplementary': supplementary, 'bursar': bursar}
        return render(request, 'finance/supplementaryincome/edit_supplementary.html', context)

    except Supplementaryincome.DoesNotExist:
        messages.error(request, 'Supplementary not found.')
        return redirect('Supplementary Income List')
# Supplementary Income views

def financeReports(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)

    context ={
        'bursar': bursar
    }
    return render(request, 'finance/financeReports.html', context)

def financeStatistics(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)

    context = {
        'bursar': bursar
    }
    return render(request, 'finance/financeStatistics.html',context)

def get_stdclass(request, stdnumber):
    try:
        student = Student.objects.get(stdnumber=stdnumber)
        classname =  student.stdclass.classname

        fees_structure = Feesstructure.objects.get(classname = classname)
        amount = fees_structure.amount

        class_data = {
            'classname': classname,
            'amount': amount
        }
        return JsonResponse(class_data)
    except Student.DoesNotExist:
        return JsonResponse({}, status=404)

def get_staff_salary(request,id):
    support_staff = Supportstaff.objects.get(pk=id)
    salary = support_staff.salary
    staff_data={
        'salary':salary,
    }
    return JsonResponse(staff_data)

def get_teacher_salary(request , id):
    teacher = Teachers.objects.get(teacherid = id)
    salary = teacher.salary
    teachersalary = {
        'salary' : salary,
    }
    return JsonResponse(teachersalary)

def get_teacher_balance(request , id):
    teacher = Teachers.objects.get(teacherid = id)
    payment = Teacherspayment.objects.get(teacherid = id)
    balance = int(teacher.salary) - int(payment.accumulatedpayment)
    print(balance)
    return JsonResponse({'balance' : balance})

def getsupportstaffbalance(request , id):
    supportstaff = Supportstaff.objects.get(supportstaffid = id)
    payment = Supportstaffpayment.objects.get(supportstaffid = id)
    balance = supportstaff.salary - payment.accumulatedamount
    return JsonResponse({'balance':balance})

def export_finance_fees_to_excel(request):
    # Fetch all the data from the Fees model
    data = Fees.objects.all().values_list(
        'paymentid', 'stdnumber__stdnumber', 'stdname', 'studentclass', 'amount', 'balance', 'modeofpayment'
    )

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write field names to the worksheet as headers
    ws.append(['Payment ID', 'Student Number', 'Student Name', 'Student Class', 'Amount', 'Balance', 'Mode of Payment'])

    # Write data to the worksheet
    for row_data in data:
        ws.append(row_data)

    # Set the column width for the date column
    ws.column_dimensions['A'].width = 15  # Adjust the width as needed
    ws.column_dimensions['B'].width = 15  # Adjust the width as needed
    ws.column_dimensions['C'].width = 15  # Adjust the width as needed
    ws.column_dimensions['D'].width = 15  # Adjust the width as needed
    ws.column_dimensions['E'].width = 15  # Adjust the width as needed
    ws.column_dimensions['F'].width = 15  # Adjust the width as needed
    ws.column_dimensions['H'].width = 15  # Adjust the width as needed

    # Set the filename and content type for the response
    filename = 'finance_fees_data.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response


def export_fees_by_class(request, class_id):
    try:
        # Fetch the selected class based on the class_id
        selected_class = Schoolclasses.objects.get(classid=class_id)

        # Fetch fees records for students in the selected class
        fees_list = Fees.objects.filter(studentclass=selected_class.classname)

        if not fees_list:
            # If there are no fees records for the selected class, return an empty response
            return HttpResponse("No fees data available for this class.")

        # Create a new workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write field names to the worksheet as headers
        ws.append(['Payment ID', 'Student Number', 'Student Name', 'Student Class', 'Amount', 'Balance', 'Mode of Payment', 'Date'])

        # Write data to the worksheet
        for fee in fees_list:
            ws.append([
                fee.paymentid,
                fee.stdnumber.stdnumber,
                fee.stdname,
                fee.studentclass,
                fee.amount,
                fee.balance,
                fee.modeofpayment,
                fee.date,
            ])

        # Set the column width for the date column
        ws.column_dimensions['A'].width = 15  # Adjust the width as needed
        ws.column_dimensions['B'].width = 15  # Adjust the width as needed
        ws.column_dimensions['C'].width = 15  # Adjust the width as needed
        ws.column_dimensions['D'].width = 15  # Adjust the width as needed
        ws.column_dimensions['E'].width = 15  # Adjust the width as needed
        ws.column_dimensions['F'].width = 15  # Adjust the width as needed
        ws.column_dimensions['H'].width = 15  # Adjust the width as needed

        # Set the filename and content type for the response
        filename = f'fees_data_{selected_class.classname}.xlsx'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Save the workbook to the response
        wb.save(response)

        return response
    except Schoolclasses.DoesNotExist:
        return HttpResponse("Class not found.", status=404)


def export_fees_structure_to_excel(request):
    data = Feesstructure.objects.all().values_list(
        'classname', 'amount'
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Class', 'Amount'])

    for row_data in data:
        ws.append(row_data)

    # Set the column width for the amount column
    ws.column_dimensions['A'].width = 15

    filename = 'fees_structuredata.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response

def export_utilities_to_excel(request):
    data = Utilities.objects.all().values_list(
        'utilitiesid', 'category', 'utilitiesdate', 'amountpaid'
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Utility ID', 'Category', 'Utilities Date', 'Amount Paid'])

    for row_data in data:
        ws.append(row_data)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    filename = 'utilities_data.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response

def export_supplementaryincome_to_excel(request):
    data = Supplementaryincome.objects.all().values_list(
        'supplementaryincomeid', 'category', 'date', 'amount'
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Supplementyaryid', 'Category', 'Date', 'Amount'])

    for row_data in data:
        ws.append(row_data)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    filename = 'supplementaryincome_data.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response

def export_teacher_payments_to_excel(request):
    data = Teacherspayment.objects.all().values_list(
        'teacherid', 'teachername', 'paymentdate', 'salary', 'amountpaid', 'balance'
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Teacher ID', 'Teacher Name', 'Payment Date', 'Salary', 'Amount Paid', 'Balance'])

    for row_data in data:
        ws.append(row_data)

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    filename = 'teacher_payments_data.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response




def export_support_staff_payments_to_excel(request):
    data = Supportstaffpayment.objects.all().values_list(
        'supportstaffid', 'staffname', 'paymentdate', 'salary', 'amountpaid', 'balance'
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append([
        'Staff ID', 'Staff Name', 'Payment Date', 'Salary', 'Amount Paid', 'Balance'
    ])

    for row_data in data:
        ws.append(row_data)

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    filename = 'support_staff_payments.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response

def export_clearedstudents_to_excel(request):
    # Query cleared students data from your Fees model (assuming balance = 0 indicates cleared students)
    cleared_students_data = Fees.objects.filter(balance=0).values_list(
        'stdnumber', 'stdname', 'studentclass'
    )

    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define column headers
    ws.append(['Student Number', 'Student Name', 'Student Class'])

    # Iterate over cleared students' data and add it to the worksheet
    for row_data in cleared_students_data:
        ws.append(row_data)

    # Set column widths (adjust as needed)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15

    # Define the filename and create an HTTP response
    filename = 'cleared_students_data.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response

def feesReport(request):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    term_data = Term.objects.get(status=1)
    terms = Term.objects.all()
    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)

    fees_list = Fees.objects.filter(term=term_data.current_term, year=term_data.current_year)
    total_amount = fees_list.aggregate(Sum('amount'))['amount__sum']
    classes = Schoolclasses.objects.all()
    context = {
        'fees_list': fees_list,
        'total_amount': total_amount,
        'classes': classes,
        'bursar': bursar,
        'terms': terms
    }
    return render(request,'finance/reports/feesreport.html',context)

def feesreport_by_class(request, class_id):
    if 'admin_id' not in request.session:
        # If the teacher is not logged in, redirect to the login page
        return redirect('financeloginpage')  # Replace 'login' with the name/url of your login view

    # Get the teacher ID from the session
    admin_id = request.session['admin_id']

    bursar = Administrators.objects.get(id=admin_id)

    classes = Schoolclasses.objects.all()
    term_data = Term.objects.get(status=1)
    
    try:
        # Fetch the selected class based on the class_id
        selected_class = Schoolclasses.objects.get(classid=class_id)
        
        # Fetch fees records for students in the selected class
        fees_list = Fees.objects.filter(studentclass=selected_class.classname, term=term_data.current_term, year=term_data.current_year)
        
        # Retrieve the class fees from the Fees model
        total_amount = fees_list.aggregate(Sum('amount'))['amount__sum']
        # classfees = classfees.amount
    except Schoolclasses.DoesNotExist:
        selected_class = None
        fees_list = []
        classfees = 0  # Default value if class is not found or fees not set
    
    return render(request, 'finance/reports/feesreport_by_class.html', {
        'fees_list': fees_list,
        'classes': classes,
        'selected_class': selected_class,
        'total_amount': total_amount,  # Pass the class fees to the template
        'bursar' : bursar
    })














